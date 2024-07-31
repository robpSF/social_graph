import numpy as np
import streamlit as st
import openpyxl
import pandas as pd
from random import randrange

# This is for the social graph drawing
from pyvis import network as net

# Constants: Row & Column indexes for the microblog social graph file
NAMES_COL = 0
NAMES_ROW = 0
HANDLES_COL = 2  # Remember columns start numbering at 0 in Python
FACTIONS_COL = 3

# Streamlit sliders for configurable parameters
st.sidebar.header("Configure Parameters")

st.sidebar.write("""
### Default Affinity
The default affinity value is used when there is no specific affinity defined between two factions. This affects the likelihood of inter-faction follow relationships.
""")
DEFAULT_AFFINITY = st.sidebar.slider("Default Affinity", min_value=0.0, max_value=1.0, value=0.1, step=0.01)

st.sidebar.write("""
### Follower Thresholds
The follower thresholds define the follower count ranges that adjust the likelihood of a persona following another within the same faction. 
- **Lower Follower Threshold**: Personas with fewer followers than this threshold have a decreased likelihood of following others within the same faction.
- **Upper Follower Threshold**: Personas with follower counts between the lower and upper thresholds have an increased likelihood of following others within the same faction.
""")
FOLLOWER_THRESHOLD_1 = st.sidebar.slider("Lower Follower Threshold", min_value=0, max_value=10000, value=1000, step=100)
FOLLOWER_THRESHOLD_2 = st.sidebar.slider("Upper Follower Threshold", min_value=1000, max_value=20000, value=10000, step=100)

st.sidebar.write("""
### Likelihood Adjustments
These adjustments define how much the likelihood of following is increased or decreased based on the follower thresholds within the same faction.
- **Likelihood Increment**: Increment added to the likelihood of following for personas within the specified follower count range.
- **Likelihood Decrement**: Decrement subtracted from the likelihood of following for personas with fewer followers than the lower threshold.
""")
LIKELIHOOD_INCREMENT = st.sidebar.slider("Likelihood Increment", min_value=0.0, max_value=1.0, value=0.2, step=0.01)
LIKELIHOOD_DECREMENT = st.sidebar.slider("Likelihood Decrement", min_value=0.0, max_value=1.0, value=0.1, step=0.01)

st.sidebar.write("""
### Display Social Graph
Toggle this option to display or hide the social graph visualization.
""")
DISPLAY_GRAPH = st.sidebar.checkbox("Display Social Graph", value=True)

# Start to create the network viz
g = net.Network(height='1000px', width='100%', bgcolor='#222222', font_color='white', directed=True)
g.set_options('''
var options = {
    "nodes": {
        "borderWidth": 1,
        "borderWidthSelected": 8,
        "color": {
            "highlight": {
                "background": "red",
                "border": "yellow"
            }
        }
    },
    "edges": {
        "color": {
            "color": "lightgray",
            "highlight": "yellow",
            "hover": "lightblue"
        },
        "width": 2,
        "hoverWidth": 3,
        "selectionWidth": 3
    },
    "physics": {
        "barnesHut": {
            "gravitationalConstant": -14900,
            "centralGravity": 2.2,
            "springLength": 110,
            "springConstant": 0.001
        },
        "minVelocity": 0.75,
        "timestep": 0.46
    }
}
''')

# This looks overall likely to follow based on reach
def whats_the_friendship(a, b, attraction_df, affinity_df):
    # Get the persona factions
    if a not in attraction_df.TwHandle.values:
        return 0, 0

    if b not in attraction_df.TwHandle.values:
        return 0, 0

    faction_a = attraction_df.loc[attraction_df.TwHandle == a, "Faction"].values[0]
    faction_b = attraction_df.loc[attraction_df.TwHandle == b, "Faction"].values[0]
    try:
        affinity_between_factions = affinity_df.loc[(affinity_df.Faction == faction_a) & (affinity_df.Other_Faction == faction_b), "Affinity"].values[0]
    except IndexError:
        affinity_between_factions = DEFAULT_AFFINITY  # Use default affinity if not found

    # FIRST pass "a is followed by b?"
    if faction_a == faction_b:  # Intra-faction probability
        likelihood_of_following = attraction_df.loc[attraction_df.TwHandle == a, "Prob4Faction"].values[0]
        affinity_between_factions = 1  # This overrides the "0" from above
        # Adjust likelihood based on follower count
        followers = attraction_df.loc[attraction_df.TwHandle == a, "TwFollowers"].values[0]
        if FOLLOWER_THRESHOLD_1 < followers < FOLLOWER_THRESHOLD_2:
            likelihood_of_following = likelihood_of_following + LIKELIHOOD_INCREMENT
        if followers < FOLLOWER_THRESHOLD_1:
            likelihood_of_following = likelihood_of_following - LIKELIHOOD_DECREMENT
    else:
        # Inter-faction probability
        likelihood_of_following = attraction_df.loc[attraction_df.TwHandle == a, "ProbOverAll"].values[0] * affinity_between_factions

    dice_roll = randrange(100) / 100
    if likelihood_of_following > dice_roll:  # If likelihood is greater than dice roll
        x = 3
    else:
        x = 0

    # SECOND pass "b is followed by a?"
    if faction_a == faction_b:  # Intra-faction probability
        likelihood_of_following = attraction_df.loc[attraction_df.TwHandle == b, "Prob4Faction"].values[0]
        affinity_between_factions = 1  # This overrides the "0" from above
        # Adjust likelihood based on follower count
        followers = attraction_df.loc[attraction_df.TwHandle == b, "TwFollowers"].values[0]
        if FOLLOWER_THRESHOLD_1 < followers < FOLLOWER_THRESHOLD_2:
            likelihood_of_following = likelihood_of_following + LIKELIHOOD_INCREMENT
        if followers < FOLLOWER_THRESHOLD_1:
            likelihood_of_following = likelihood_of_following - LIKELIHOOD_DECREMENT
    else:
        # Inter-faction probability
        likelihood_of_following = attraction_df.loc[attraction_df.TwHandle == b, "ProbOverAll"].values[0] * affinity_between_factions

    dice_roll = randrange(100) / 100
    if likelihood_of_following > dice_roll:  # If likelihood is greater than dice roll
        y = 1
    else:
        y = 0

    # If they follow each other then it's a 2!
    if x == 1 and y == 3:
        x = 2
        y = 2

    return x, y


# Input fields for file uploads
st.title("Microblog Social Graph")
st.write("""
## Explanation of Follower Thresholds

### Intra-Faction Probability
When two personas are from the same faction, the follower thresholds (FOLLOWER_THRESHOLD_1 and FOLLOWER_THRESHOLD_2) are used to adjust the likelihood of following.
If a persona's follower count is within a certain range, the likelihood of them following another persona within the same faction is either increased or decreased.

### Inter-Faction Probability
When two personas are from different factions, the follower thresholds are not directly applied. Instead, the likelihood of following is adjusted based on the affinity between the factions.
The overall probability of following is determined by the affinity between the factions and the ProbOverAll value for the personas.
""")

persona_details = st.file_uploader("Upload persona_details.xlsx", type=["xlsx"])
social_graph = st.file_uploader("Upload Microblog_social_graph.xlsx", type=["xlsx"])

if persona_details and social_graph:
    df = pd.read_excel(persona_details)
    df2 = df[["Faction", "TwFollowers", "TwHandle", "TwBio"]]
    df2 = df2.sort_values(["Faction", "TwFollowers"], ascending=False).groupby("Faction").head(1000)
    st.write(df2)
    df2["Prob4Faction"] = df2["TwFollowers"] / df2.groupby("Faction")["TwFollowers"].transform("sum")
    df2["ProbOverAll"] = df2["TwFollowers"] / df2["TwFollowers"].sum()
    st.write(df2)
    attraction_df = df2

    source_wb = openpyxl.load_workbook(social_graph)
    social_graph_sheet = source_wb.active
    max_rows = social_graph_sheet.max_row

    handles = [social_graph_sheet.cell(row=i, column=HANDLES_COL + 1).value for i in range(2, max_rows + 1)]
    handles = [handle for handle in handles if handle]  # Remove None values

    factions = [social_graph_sheet.cell(row=i, column=FACTIONS_COL + 1).value for i in range(2, max_rows + 1)]
    factions = list(dict.fromkeys(factions))
    if "Faction" in factions:
        factions.remove("Faction")
    if "" in factions:
        factions.remove("")
    st.write(factions)

    factions_df = pd.DataFrame(factions, columns=["Faction"])
    factions_df["Other_Faction"] = ""
    factions_df["Affinity"] = 0
    st.write(factions_df)

    st.write("Please edit the faction affinity below and then save:")
    affinity_file = st.file_uploader("Upload Faction affinity file", type=["xlsx"])
    if affinity_file:
        affinity_df = pd.read_excel(affinity_file)
        affinity_df = affinity_df[["Faction", "Other_Faction", "Affinity"]]
        st.table(affinity_df)

        # Precompute affinity lookup
        affinity_dict = affinity_df.set_index(['Faction', 'Other_Faction'])['Affinity'].to_dict()

        # Precompute likelihoods
        df2['Likelihood'] = df2.apply(lambda row: calculate_likelihood(row['TwHandle'], row['TwFollowers'], row['Prob4Faction'], row['ProbOverAll']), axis=1)

        # Progress bar
        progress_text = st.empty()
        progress_bar = st.progress(0)

        total_steps = len(handles) + (len(handles) * (len(handles) - 1)) // 2
        step = 0

        for i in range(1, max_rows-1):
            persona = handles[i]
            bio = ""
            faction = ""
            try:
                bio = df2.loc[df2.TwHandle == persona, "TwBio"].values[0]
                faction = df2.loc[df2.TwHandle == persona, "Faction"].values[0]
            except IndexError:
                bio = " "

            g.add_node(persona, title=f"({persona})[{faction}] {bio}")

            step += 1
            progress_percentage = step / total_steps
            progress_text.text(f"Processing nodes: {step} / {len(handles)}")
            progress_bar.progress(progress_percentage)

        def add_edges(df):
            edges = []
            for i in range(len(df)):
                for j in range(i + 1, len(df)):
                    followed = df.iloc[i]
                    follower = df.iloc[j]

                    friend_value_x, friend_value_y = whats_the_friendship(followed['TwHandle'], follower['TwHandle'], attraction_df, affinity_df)

                    if friend_value_x > 0:
                        edges.append((follower['TwHandle'], followed['TwHandle']))
                    if friend_value_y > 0:
                        edges.append((followed['TwHandle'], follower['TwHandle']))

            return edges

        edges = add_edges(df2)

        for follower, followed in edges:
            g.add_edge(follower, followed)

        output_path = 'social_OUTPUT.xlsx'
        source_wb.save(output_path)
        st.write("Social graph processing complete. Download the updated social graph:")
        st.download_button(label="Download social_OUTPUT.xlsx", data=open(output_path, 'rb').read(), file_name=output_path)

        if DISPLAY_GRAPH:
            g.save_graph('networkviz.html')
            HtmlFile = open("networkviz.html", 'r', encoding='utf-8')
            source_code = HtmlFile.read()
            st.components.v1.html(source_code, height=1000, width=1000)

    st.header("All Done!")
